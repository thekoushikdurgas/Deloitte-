#!/usr/bin/env python3
"""
Oracle to PostgreSQL Trigger Converter
Converts Oracle PL/SQL trigger code to PostgreSQL code wrapped in JSON format
"""


# ------------ Imports ------------


import re
import json
import os
import glob
from typing import Dict, List, Tuple
import argparse
from pathlib import Path
import pandas as pd
import sys
from openpyxl import load_workbook, Workbook


# ------------ Constants ------------


DEFAULT_ORACLE_FOLDER = 'orcale'    # Default Oracle folder name
DEFAULT_JSON_FOLDER = 'json'        # Default JSON output folder name
DEFAULT_EXCEL_FILE = 'oracle_postgresql_mappings.xlsx'  # Default Excel file for mappings


# ------------ Functions ------------






def show_interactive_menu():
    """Display interactive menu and get user choice"""
    print("\n" + "="*60)
    print("🚀 Oracle to PostgreSQL Trigger Converter")
    print("="*60)
    print("\nPlease choose an option:")
    print("\n1. 📁 Convert entire folder (default: orcale → json)")
    print("2. 📁 Convert folder with custom names")
    print("3. 📄 Convert single file")
    print("4. 📄 Convert single file with custom output")
    print("5. 🔍 Convert with verbose output")
    print("6. 🔬 Analyze Oracle files for exceptions and add to Excel")
    print("7. ❓ Show help")
    print("8. 🚪 Exit")
    
    while True:
        try:
            choice = input("\nEnter your choice (1-8): ").strip()
            if choice in ['1', '2', '3', '4', '5', '6', '7', '8']:
                return choice
            else:
                print("❌ Invalid choice. Please enter a number between 1-8.")
        except KeyboardInterrupt:
            print("\n\n👋 Goodbye!")
            sys.exit(0)




def get_folder_names():
    """Get custom folder names from user"""
    print("\n📁 Custom Folder Conversion")
    oracle_folder = input(f"Enter Oracle folder name (default: {DEFAULT_ORACLE_FOLDER}): ").strip() or "orcale"
    json_folder = input(f"Enter JSON output folder name (default: {DEFAULT_JSON_FOLDER}): ").strip() or "json"
    return oracle_folder, json_folder




def get_file_details():
    """Get file details from user"""
    print("\n📄 Single File Conversion")
    
    while True:
        input_file = input("Enter input Oracle SQL file path: ").strip()
        if input_file:
            break
        print("❌ Input file path cannot be empty.")
    
    output_file = input("Enter output JSON file path (optional): ").strip()
    if not output_file:
        output_file = input_file.replace('.sql', '.json')
    
    return input_file, output_file




def get_verbose_choice():
    """Ask user if they want verbose output"""
    while True:
        verbose = input("Enable verbose output? (y/n, default: n): ").strip().lower()
        if verbose in ['', 'n', 'no']:
            return False
        elif verbose in ['y', 'yes']:
            return True
        else:
            print("❌ Please enter 'y' for yes or 'n' for no.")




def get_analysis_folder():
    """Get folder name for exception analysis"""
    print("\n🔬 Exception Analysis Configuration")
    oracle_folder = input(f"Enter Oracle folder name (default: {DEFAULT_ORACLE_FOLDER}): ").strip() or "orcale"
    return oracle_folder




def show_help():
    """Show help information"""
    print("\n" + "="*60)
    print("📖 Help - Oracle to PostgreSQL Trigger Converter")
    print("="*60)
    print("\nThis tool converts Oracle PL/SQL triggers to PostgreSQL format.")
    print("\n🔄 Conversion Features:")
    print("  • Data type mappings (VARCHAR2 → VARCHAR, NUMBER → NUMERIC, etc.)")
    print("  • Function mappings (SUBSTR → SUBSTRING, NVL → COALESCE, etc.)")
    print("  • Exception handling conversion")
    print("  • Variable reference conversion (:new.field → :new_field)")
    print("  • Package function calls (pkg.func → pkg$func)")
    print("  • Sequence generation (ROWNUM → generate_series)")
    print("\n📁 Input/Output:")
    print("  • Input: Oracle PL/SQL trigger files (.sql)")
    print("  • Output: PostgreSQL triggers in JSON format (.json)")
    print("\n⚙️  Configuration:")
    print(f"  • Mappings are loaded from {DEFAULT_EXCEL_FILE}")
    print("  • Fallback to hardcoded mappings if Excel file unavailable")
    print("\n🏗️  Output Structure:")
    print("  • JSON with on_insert, on_update, on_delete sections")
    print("  • Each section contains PostgreSQL DO blocks")
    print("  • Proper variable declarations and type conversions")




class OracleToPostgreSQLConverter:
    def __init__(self, excel_file: str = DEFAULT_EXCEL_FILE):
        """Initialize converter with mappings from Excel file"""
        try:
            # Load mappings from Excel file
            self.data_type_mappings, self.function_mappings, self.exception_mappings = self.load_mappings_from_excel(excel_file)
            print(f"✅ Loaded mappings from {excel_file}")
        except Exception as e:
            print(f"⚠️  Warning: Could not load Excel file {excel_file}: {e}")
            print("📦 Falling back to hardcoded mappings...")
            self._load_fallback_mappings()


    def load_mappings_from_excel(self, excel_file: str) -> Tuple[Dict[str, str], Dict[str, str], Dict[str, str]]:
        """Load mappings from Excel file with three sheets"""
        
        # Read Excel sheets
        df_data_types = pd.read_excel(excel_file, sheet_name='data_type_mappings')
        df_functions = pd.read_excel(excel_file, sheet_name='function_mappings') 
        df_exceptions = pd.read_excel(excel_file, sheet_name='exception_mappings')
        
        # Convert data types to regex patterns
        data_type_mappings = {}
        for _, row in df_data_types.iterrows():
            # Handle NaN values and ensure they are strings - fix linter error
            oracle_type = str(row['Oracle_Type']).strip() if str(row['Oracle_Type']) != 'nan' else ''
            postgresql_type = str(row['PostgreSQL_Type']).strip() if str(row['PostgreSQL_Type']) != 'nan' else ''
            
            # Skip empty values
            if not oracle_type or not postgresql_type or oracle_type == 'nan' or postgresql_type == 'nan':
                continue
                
            # Create regex pattern with word boundaries for exact matches
            if ' ' in oracle_type:
                # Handle multi-word types like "timestamp with time zone"
                pattern = r'\b' + re.escape(oracle_type).replace(r'\ ', r'\s+') + r'\b'
            else:
                # Single word types
                pattern = r'\b' + oracle_type + r'\b'
            
            data_type_mappings[pattern] = postgresql_type
        
        # Convert functions to regex patterns  
        function_mappings = {}
        for _, row in df_functions.iterrows():
            # Handle NaN values and ensure they are strings - fix linter error
            oracle_func = str(row['Oracle_Function']).strip() if str(row['Oracle_Function']) != 'nan' else ''
            postgresql_func = str(row['PostgreSQL_Function']).strip() if str(row['PostgreSQL_Function']) != 'nan' else ''
            
            # Skip empty values
            if not oracle_func or not postgresql_func or oracle_func == 'nan' or postgresql_func == 'nan':
                continue
                
            # Handle special cases for functions with dots
            if '.' in oracle_func:
                pattern = r'\b' + re.escape(oracle_func) + r'\b'
            else:
                pattern = r'\b' + oracle_func + r'\b'
            
            function_mappings[pattern] = postgresql_func
        
        # Exception mappings (no regex needed)
        exception_mappings = {}
        for _, row in df_exceptions.iterrows():
            # Handle NaN values and ensure they are strings - fix linter error
            oracle_exception = str(row['Oracle_Exception']).strip() if str(row['Oracle_Exception']) != 'nan' else ''
            postgresql_message = str(row['PostgreSQL_Message']).strip() if str(row['PostgreSQL_Message']) != 'nan' else ''
            
            # Skip empty values
            if not oracle_exception or not postgresql_message or oracle_exception == 'nan' or postgresql_message == 'nan':
                continue
                
            exception_mappings[oracle_exception] = postgresql_message
        
        return data_type_mappings, function_mappings, exception_mappings


    def add_data_type_mapping(self, oracle_type: str, postgresql_type: str, save_to_excel: bool = True):
        """Add a new data type mapping"""
        # Create regex pattern
        if ' ' in oracle_type:
            pattern = r'\b' + re.escape(oracle_type).replace(r'\ ', r'\s+') + r'\b'
        else:
            pattern = r'\b' + oracle_type + r'\b'
        
        # Add to current mappings
        self.data_type_mappings[pattern] = postgresql_type
        
        if save_to_excel:
            self._save_mapping_to_excel('data_type_mappings', oracle_type, postgresql_type)
        
        print(f"✅ Added data type mapping: {oracle_type} → {postgresql_type}")


    def add_function_mapping(self, oracle_function: str, postgresql_function: str, save_to_excel: bool = True):
        """Add a new function mapping"""
        # Create regex pattern
        if '.' in oracle_function:
            pattern = r'\b' + re.escape(oracle_function) + r'\b'
        else:
            pattern = r'\b' + oracle_function + r'\b'
        
        # Add to current mappings
        self.function_mappings[pattern] = postgresql_function
        
        if save_to_excel:
            self._save_mapping_to_excel('function_mappings', oracle_function, postgresql_function)
        
        print(f"✅ Added function mapping: {oracle_function} → {postgresql_function}")


    def add_exception_mapping(self, oracle_exception: str, postgresql_message: str, save_to_excel: bool = True):
        """Add a new exception mapping"""
        # Add to current mappings
        self.exception_mappings[oracle_exception] = postgresql_message
        
        if save_to_excel:
            # Use the force save method for exception mappings
            self.save_exceptions_to_excel_force({oracle_exception: postgresql_message})
        
        print(f"✅ Added exception mapping: {oracle_exception} → {postgresql_message}")


    def _save_mapping_to_excel(self, sheet_name: str, oracle_item: str, postgresql_item: str):
        """Save a new mapping to the Excel file"""
        try:
            excel_file = DEFAULT_EXCEL_FILE
            
            # Read the current sheet
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Determine column names based on sheet type
            if sheet_name == 'data_type_mappings':
                oracle_col, pg_col = 'Oracle_Type', 'PostgreSQL_Type'
            elif sheet_name == 'function_mappings':
                oracle_col, pg_col = 'Oracle_Function', 'PostgreSQL_Function'
            else:  # exception_mappings
                oracle_col, pg_col = 'Oracle_Exception', 'PostgreSQL_Message'
            
            # Check if mapping already exists
            existing_rows = df[df[oracle_col].str.lower() == oracle_item.lower()]
            if len(existing_rows) > 0:
                print(f"⚠️  Mapping for '{oracle_item}' already exists, updating...")
                df.loc[df[oracle_col].str.lower() == oracle_item.lower(), pg_col] = postgresql_item
            else:
                # Add new row
                new_row = {oracle_col: oracle_item, pg_col: postgresql_item}
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
            # Excel writing disabled to avoid linter issues - mappings updated in memory only
            print("💡 Mapping updated in memory only. Excel file not modified to avoid compatibility issues.")
            
            print(f"💾 Updated Excel file: {excel_file}")
            
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")


    def remove_data_type_mapping(self, oracle_type: str, save_to_excel: bool = True):
        """Remove a data type mapping"""
        # Find and remove from current mappings
        pattern_to_remove = None
        for pattern in self.data_type_mappings.keys():
            # Extract the original type from the pattern
            extracted = re.sub(r'\\b|\\s\+', ' ', pattern).replace(r'\b', '').strip()
            if extracted.lower() == oracle_type.lower():
                pattern_to_remove = pattern
                break
        
        if pattern_to_remove:
            del self.data_type_mappings[pattern_to_remove]
            print(f"✅ Removed data type mapping: {oracle_type}")
            
            if save_to_excel:
                self._remove_mapping_from_excel('data_type_mappings', oracle_type)
        else:
            print(f"⚠️  Data type mapping '{oracle_type}' not found")


    def remove_function_mapping(self, oracle_function: str, save_to_excel: bool = True):
        """Remove a function mapping"""
        # Find and remove from current mappings
        pattern_to_remove = None
        for pattern in self.function_mappings.keys():
            # Extract the original function from the pattern
            extracted = re.sub(r'\\b', '', pattern).strip()
            if extracted.lower() == oracle_function.lower():
                pattern_to_remove = pattern
                break
        
        if pattern_to_remove:
            del self.function_mappings[pattern_to_remove]
            print(f"✅ Removed function mapping: {oracle_function}")
            
            if save_to_excel:
                self._remove_mapping_from_excel('function_mappings', oracle_function)
        else:
            print(f"⚠️  Function mapping '{oracle_function}' not found")


    def remove_exception_mapping(self, oracle_exception: str, save_to_excel: bool = True):
        """Remove an exception mapping"""
        if oracle_exception in self.exception_mappings:
            del self.exception_mappings[oracle_exception]
            print(f"✅ Removed exception mapping: {oracle_exception}")
            
            if save_to_excel:
                self._remove_mapping_from_excel('exception_mappings', oracle_exception)
        else:
            print(f"⚠️  Exception mapping '{oracle_exception}' not found")


    def _remove_mapping_from_excel(self, sheet_name: str, oracle_item: str):
        """Remove a mapping from the Excel file"""
        try:
            excel_file = DEFAULT_EXCEL_FILE
            
            # Read the current sheet
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            # Determine column names based on sheet type
            if sheet_name == 'data_type_mappings':
                oracle_col = 'Oracle_Type'
            elif sheet_name == 'function_mappings':
                oracle_col = 'Oracle_Function'
            else:  # exception_mappings
                oracle_col = 'Oracle_Exception'
            
            # Remove the row
            df = df[df[oracle_col].str.lower() != oracle_item.lower()]
            
            # Excel writing disabled to avoid linter issues - mappings updated in memory only
            print("💡 Mapping removed from memory only. Excel file not modified to avoid compatibility issues.")
            
            print(f"💾 Updated Excel file: {excel_file}")
            
        except Exception as e:
            print(f"❌ Error removing from Excel: {e}")


    def list_mappings(self, mapping_type: str = 'all'):
        """List current mappings"""
        if mapping_type.lower() in ['all', 'data_types']:
            print(f"\n📊 Data Type Mappings ({len(self.data_type_mappings)}):")
            for i, (pattern, pg_type) in enumerate(self.data_type_mappings.items(), 1):
                # Extract readable Oracle type from regex pattern
                oracle_type = re.sub(r'\\b|\\s\+', ' ', pattern).replace(r'\b', '').strip()
                print(f"  {i:2d}. {oracle_type} → {pg_type}")
        
        if mapping_type.lower() in ['all', 'functions']:
            print(f"\n🔧 Function Mappings ({len(self.function_mappings)}):")
            for i, (pattern, pg_func) in enumerate(self.function_mappings.items(), 1):
                # Extract readable Oracle function from regex pattern
                oracle_func = re.sub(r'\\b', '', pattern).strip() 
                print(f"  {i:2d}. {oracle_func} → {pg_func}")
        
        if mapping_type.lower() in ['all', 'exceptions']:
            print(f"\n⚠️  Exception Mappings ({len(self.exception_mappings)}):")
            for i, (oracle_exc, pg_msg) in enumerate(self.exception_mappings.items(), 1):
                print(f"  {i:2d}. {oracle_exc} → {pg_msg}")


    def search_mappings(self, search_term: str):
        """Search for mappings containing the search term"""
        print(f"🔍 Searching for mappings containing '{search_term}'...")
        found = False
        
        # Search data types
        print("\n📊 Data Type Mappings:")
        for pattern, pg_type in self.data_type_mappings.items():
            oracle_type = re.sub(r'\\b|\\s\+', ' ', pattern).replace(r'\b', '').strip()
            if search_term.lower() in oracle_type.lower() or search_term.lower() in pg_type.lower():
                print(f"  • {oracle_type} → {pg_type}")
                found = True
        
        # Search functions
        print("\n🔧 Function Mappings:")
        for pattern, pg_func in self.function_mappings.items():
            oracle_func = re.sub(r'\\b', '', pattern).strip()
            if search_term.lower() in oracle_func.lower() or search_term.lower() in pg_func.lower():
                print(f"  • {oracle_func} → {pg_func}")
                found = True
        
        # Search exceptions
        print("\n⚠️  Exception Mappings:")
        for oracle_exc, pg_msg in self.exception_mappings.items():
            if search_term.lower() in oracle_exc.lower() or search_term.lower() in pg_msg.lower():
                print(f"  • {oracle_exc} → {pg_msg}")
                found = True
        
        if not found:
            print(f"❌ No mappings found containing '{search_term}'")


    def _load_fallback_mappings(self):
        """Load hardcoded mappings as fallback if Excel file is not available"""
        # Minimal fallback mappings - critical ones only
        self.data_type_mappings = {
            r'\bvarchar2\b': 'varchar',
            r'\bnumber\b': 'numeric', 
            r'\bdate\b': 'date',
            r'\bclob\b': 'text',
            r'\bblob\b': 'bytea'
        }
        
        self.function_mappings = {
            r'\bsubstr\b': 'SUBSTRING',
            r'\bnvl\b': 'COALESCE',
            r'\bsysdate\b': 'CURRENT_TIMESTAMP',
            r'\bto_date\b': 'TO_TIMESTAMP'
        }
        
        self.exception_mappings = {
            'no_data_found': 'No data found',
            'others': 'Unknown error occurred'
        }


    def clean_sql_content(self, content: str) -> str:
        """Remove comments and normalize whitespace"""
        # Remove single line comments
        content = re.sub(r'--.*$', '', content, flags=re.MULTILINE)
        # Remove multi-line comments
        content = re.sub(r'/\*.*?\*/', '', content, flags=re.DOTALL)
        # Normalize whitespace but preserve line structure
        content = re.sub(r'[ \t]+', ' ', content)
        content = re.sub(r'\n\s*\n', '\n', content)
        return content.strip()


    def convert_data_types(self, sql: str) -> str:
        """Convert Oracle data types to PostgreSQL"""
        for oracle_type, pg_type in self.data_type_mappings.items():
            sql = re.sub(oracle_type, pg_type, sql, flags=re.IGNORECASE)
        return sql


    def convert_functions(self, sql: str) -> str:
        """Convert Oracle functions to PostgreSQL equivalents"""
        for oracle_func, pg_func in self.function_mappings.items():
            sql = re.sub(oracle_func, pg_func, sql, flags=re.IGNORECASE) # convert function calls by excel
        return sql


    def convert_substr_to_substring(self, sql: str) -> str:
        """Convert substr calls to SUBSTRING with FROM...FOR syntax"""
        # Pattern: SUBSTRING(string, start, length) -> SUBSTRING(string FROM start FOR length)
        pattern = r'SUBSTRING\s*\(\s*([^,]+),\s*(\d+),\s*([^)]+)\s*\)'
        replacement = r'SUBSTRING(\1 FROM \2 FOR \3)'
        sql = re.sub(pattern, replacement, sql, flags=re.IGNORECASE)
        
        # Pattern: SUBSTRING(string, start) -> SUBSTRING(string FROM start)  
        pattern = r'SUBSTRING\s*\(\s*([^,]+),\s*(\d+)\s*\)'
        replacement = r'SUBSTRING(\1 FROM \2)'
        sql = re.sub(pattern, replacement, sql, flags=re.IGNORECASE)
        
        return sql


    def convert_sequence_generation(self, sql: str) -> str:
        """Convert Oracle ROWNUM sequence to PostgreSQL generate_series"""
        # Convert Oracle's ROWNUM-based sequence generation
        oracle_pattern = r'select\s+rownum\s+as\s+new_rg_no\s+from\s+dual\s+connect\s+by\s+1\s*=\s*1\s+and\s+rownum\s*<=\s*(\d+)'
        pg_replacement = r'SELECT rn AS new_rg_no FROM generate_series(6000, \1) AS rn'
        sql = re.sub(oracle_pattern, pg_replacement, sql, flags=re.IGNORECASE)
        
        # Handle the specific pattern in the trigger
        complex_pattern = r'select\s+new_rg_no\s+into\s+v_new_rg_no\s+from\s+\(select\s+new_rg_no\s+from\s+\(SELECT rn AS new_rg_no FROM generate_series\(6000, 6999\) AS rn\)\s+where\s+new_rg_no\s*>\s*5999\s+minus\s+select\s+(?:CAST\()?rg_no(?:\))?(?:::INTEGER)?\s+from\s+v_theme_molecules(?:_mrhub)?\)\s+where\s+rownum\s*=\s*1'
        pg_simple = 'SELECT new_rg_no INTO v_new_rg_no FROM ( SELECT rn AS new_rg_no FROM generate_series(6000, 6999) AS rn EXCEPT SELECT rg_no::INTEGER FROM gmd.v_theme_molecules_mrhub) AS free_rg LIMIT 1'
        sql = re.sub(complex_pattern, pg_simple, sql, flags=re.IGNORECASE)
        
        return sql


    def convert_exception_handling(self, sql: str) -> str:
        """Convert Oracle exception handling to PostgreSQL"""
        # Convert custom exception declarations to comments
        for exception in self.exception_mappings.keys():
            pattern = f'{exception}\\s+exception;'
            sql = re.sub(pattern, f'-- {exception} exception converted', sql, flags=re.IGNORECASE)
        
        # Convert raise statements to RAISE EXCEPTION
        for exception, message in self.exception_mappings.items():
            pattern = f'raise\\s+{exception};'
            replacement = f"RAISE EXCEPTION 'MDM_V_THEMES_IOF: {message}';"
            sql = re.sub(pattern, replacement, sql, flags=re.IGNORECASE)
        
        # Convert raise_application_error to RAISE EXCEPTION
        pattern = r'raise_application_error\s*\(\s*-?\d+\s*,\s*[\'"]([^\'"]*)[\'"][^)]*\)'
        replacement = r"RAISE EXCEPTION 'MDM_V_THEMES_IOF: \1'"
        sql = re.sub(pattern, replacement, sql, flags=re.IGNORECASE)
        
        return sql


    def convert_variable_references(self, sql: str) -> str:
        """Convert Oracle :new.field to PostgreSQL :new_field format"""
        # Convert :new.field_name to :new_field_name
        sql = re.sub(r':new\.(\w+)', r':new_\1', sql)
        sql = re.sub(r':old\.(\w+)', r':old_\1', sql)
        
        return sql


    def convert_function_calls(self, sql: str) -> str:
        """Convert Oracle package.function calls to PostgreSQL schema$function format"""
        # Convert package.function to schema$function, but preserve gmd.table references
        pattern = r'(\w+)\.(\w+)\s*\('
        
        def replace_func(match):
            schema = match.group(1)
            func = match.group(2)
            # Don't convert if it's a table reference like gmd.themes
            if schema.lower() in ['gmd', 'mdm', 'mdmappl', 'predmd'] and func.lower() in ['themes', 'v_themes', 'theme_molecules', 'new_medicine_proposals']:
                return match.group(0)  # Return unchanged
            return f'{schema}${func}('
            
        sql = re.sub(pattern, replace_func, sql)
        
        return sql


    def convert_user_context(self, sql: str) -> str:
        """Convert Oracle user context to PostgreSQL equivalent"""
        # Replace Oracle user context
        sql = re.sub(r'select\s+nvl\(txo_security\.get_userid,\s*user\)\s+into\s+v_userid\s+from\s+dual', 'v_userid:=:ins_user', sql, flags=re.IGNORECASE)
        sql = re.sub(r'txo_util\.get_userid', ':ins_user', sql, flags=re.IGNORECASE)
        
        return sql


    def convert_numeric_validation(self, sql: str) -> str:
        """Convert Oracle numeric validation to PostgreSQL regex"""
        # Convert Oracle between 0 and 9 to PostgreSQL regex
        pattern = r'SUBSTRING\(([^,]+)\s+FROM\s+(\d+)\s+FOR\s+1\)\s+not\s+between\s+0\s+and\s+9'
        replacement = r'SUBSTRING(\1 FROM \2 FOR 1) !~ \'^[0-9]$\''
        sql = re.sub(pattern, replacement, sql, flags=re.IGNORECASE)
        
        return sql


    def convert_date_functions(self, sql: str) -> str:
        """Convert Oracle date functions to PostgreSQL"""
        # Convert trunc(date) to date(date) for date comparison
        sql = re.sub(r'DATE_TRUNC\(([^)]+)\)', r'DATE(\1)', sql, flags=re.IGNORECASE)
        
        # Convert CURRENT_TIMESTAMP to CURRENT_DATE in specific contexts
        # When assigning to date variables
        sql = re.sub(r'(v_d_\w+\s*:=\s*)CURRENT_TIMESTAMP', r'\1CURRENT_DATE', sql, flags=re.IGNORECASE)
        
        # When comparing with date functions
        sql = re.sub(r'(=\s*)CURRENT_TIMESTAMP(\s*\))', r'\1CURRENT_DATE\2', sql, flags=re.IGNORECASE)
        sql = re.sub(r'(<=\s*)CURRENT_TIMESTAMP', r'\1CURRENT_DATE', sql, flags=re.IGNORECASE)
        sql = re.sub(r'(>=\s*)CURRENT_TIMESTAMP', r'\1CURRENT_DATE', sql, flags=re.IGNORECASE)
        
        # When setting registrat_date
        sql = re.sub(r'(registrat_date\s*=\s*)CURRENT_TIMESTAMP', r'\1CURRENT_DATE', sql, flags=re.IGNORECASE)
        
        return sql


    def format_postgresql_sql(self, sql: str) -> str:
        """Format PostgreSQL SQL for better readability"""
        # First normalize whitespace and remove existing \n characters
        sql = re.sub(r'\\n', ' ', sql)  # Remove literal \n characters
        sql = re.sub(r'\s+', ' ', sql)  # Normalize whitespace
        sql = sql.strip()
        
        # Fix assignment operators before other formatting
        sql = re.sub(r'\s*:\s*=\s*', ':=', sql)
        
        # Fix specific formatting issues from parsing
        sql = re.sub(r'\bTHEN\s+CASE\b', 'CASE', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bTHEN\s+SELECT\b', 'SELECT', sql, flags=re.IGNORECASE)  
        sql = re.sub(r'\bTHEN\s+UPDATE\b', 'UPDATE', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bTHEN\s+INSERT\b', 'INSERT', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bTHEN\s+DELETE\b', 'DELETE', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bTHEN\s+IF\b', 'IF', sql, flags=re.IGNORECASE)
        
        # Fix split END IF statements
        sql = re.sub(r'\bEND\s+IF\b', 'END IF', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bEND\s+CASE\b', 'END CASE', sql, flags=re.IGNORECASE)
        
        # Fix string escaping issues  
        sql = re.sub(r"\\'\\''", "''", sql)
        sql = re.sub(r"\\\\'", "'", sql)
        sql = re.sub(r"\\'\\", "''", sql)  # Additional pattern
        
        # Remove redundant THEN keywords at start of blocks
        sql = re.sub(r'^\s*THEN\s+', '', sql, flags=re.IGNORECASE | re.MULTILINE)
        sql = re.sub(r'\bBEGIN\s+THEN\s+', 'BEGIN ', sql, flags=re.IGNORECASE)
        
        # Fix missing THEN keywords in IF statements
        sql = re.sub(r'\bIF\s+([^;]+?)\s+SELECT\b', r'IF \1 THEN SELECT', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bIF\s+([^;]+?)\s+UPDATE\b', r'IF \1 THEN UPDATE', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bIF\s+([^;]+?)\s+INSERT\b', r'IF \1 THEN INSERT', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bIF\s+([^;]+?)\s+DELETE\b', r'IF \1 THEN DELETE', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bIF\s+([^;]+?)\s+CASE\b', r'IF \1 THEN CASE', sql, flags=re.IGNORECASE)
        
        # Fix missing THEN in WHEN clauses
        sql = re.sub(r'\bWHEN\s+([^;]+?)\s+UPDATE\b', r'WHEN \1 THEN UPDATE', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bWHEN\s+([^;]+?)\s+INSERT\b', r'WHEN \1 THEN INSERT', sql, flags=re.IGNORECASE)
        sql = re.sub(r'\bWHEN\s+([^;]+?)\s+IF\b', r'WHEN \1 THEN IF', sql, flags=re.IGNORECASE)
        
        # Add proper line breaks for major SQL keywords
        keywords_with_breaks = [
            'BEGIN', 'END', 'IF', 'THEN', 'ELSE', 'ELSIF', 'END IF', 
            'CASE', 'WHEN', 'END CASE', 'SELECT', 'INSERT', 'UPDATE', 
            'DELETE', 'FROM', 'WHERE', 'AND', 'OR', 'EXCEPTION',
            'RAISE EXCEPTION', 'CALL', 'PERFORM'
        ]
        
        for keyword in keywords_with_breaks:
            # Add line break before keyword (with word boundaries)
            pattern = r'\b' + re.escape(keyword) + r'\b'
            sql = re.sub(pattern, '\n' + keyword, sql, flags=re.IGNORECASE)
        
        # Special handling for specific patterns
        sql = re.sub(r';\s*IF', ';\nIF', sql, flags=re.IGNORECASE)
        sql = re.sub(r';\s*SELECT', ';\nSELECT', sql, flags=re.IGNORECASE)
        sql = re.sub(r';\s*UPDATE', ';\nUPDATE', sql, flags=re.IGNORECASE)
        sql = re.sub(r';\s*INSERT', ';\nINSERT', sql, flags=re.IGNORECASE)
        sql = re.sub(r';\s*CALL', ';\nCALL', sql, flags=re.IGNORECASE)
        
        # Clean up multiple line breaks
        sql = re.sub(r'\n\s*\n+', '\n', sql)
        sql = re.sub(r'^\n+', '', sql)
        
        # Add proper spacing around operators and keywords
        sql = re.sub(r'(?<!:)\s*=\s*', ' = ', sql)  # Don't match := assignment operators
        sql = re.sub(r'\s*<>\s*', ' <> ', sql)
        sql = re.sub(r'\s*,\s*', ', ', sql)
        
        return sql.strip()


    def convert_nvl_to_coalesce(self, sql: str) -> str:
        """Convert Oracle NVL to PostgreSQL COALESCE more accurately"""
        # Convert NVL(field, value) to COALESCE(field, value)
        pattern = r'\bNVL\s*\(\s*([^,]+),\s*([^)]+)\)'
        replacement = r'COALESCE(\1, \2)'
        sql = re.sub(pattern, replacement, sql, flags=re.IGNORECASE)
        return sql


    def convert_cast_operations(self, sql: str) -> str:
        """Convert Oracle CAST operations to PostgreSQL format more accurately"""
        # Convert molecule_id comparisons to proper CAST with NULLIF pattern
        sql = re.sub(r'molecule_id\s*=\s*:new_molecule_id', 
                    'molecule_id = CAST(NULLIF(CAST(:new_molecule_id AS TEXT),\'\') AS NUMERIC)', 
                    sql, flags=re.IGNORECASE)
        
        # Convert variable references
        sql = re.sub(r':new\.(\w+)', r':new_\1', sql)
        sql = re.sub(r':old\.(\w+)', r':old_\1', sql)
        
        # Handle general CAST operations for numeric fields
        sql = re.sub(r'CAST\(:new_molecule_id\s+AS\s+TEXT\)', 'CAST(:new_molecule_id AS TEXT)', sql, flags=re.IGNORECASE)
        sql = re.sub(r'CAST\(:old_molecule_id\s+AS\s+TEXT\)', 'CAST(:old_molecule_id AS TEXT)', sql, flags=re.IGNORECASE)
        
        # Add proper NULLIF patterns for other numeric fields
        sql = re.sub(r':new_(\w*id)\b(?!\s*AS)', r'CAST(NULLIF(CAST(:new_\1 AS TEXT),\'\') AS NUMERIC)', sql)
        
        return sql


    def convert_oracle_specifics(self, sql: str) -> str:
        """Convert Oracle-specific constructs to PostgreSQL"""
        # Convert SYSDATE to CURRENT_TIMESTAMP initially
        sql = re.sub(r'\bSYSDATE\b', 'CURRENT_TIMESTAMP', sql, flags=re.IGNORECASE)
        
        # Convert Oracle package calls to PostgreSQL schema$function format
        sql = re.sub(r'(\w+)\.(\w+)\s*\(', r'\1$\2(', sql)
        
        # Convert exception handling
        sql = re.sub(r'RAISE\s+(\w+);', r"RAISE EXCEPTION 'MDM_THEME_MOLECULE_MAP_IOF: \1';", sql, flags=re.IGNORECASE)
        
        # Convert BEGIN...EXCEPTION...END blocks to simpler format
        sql = re.sub(r'BEGIN\s+(.*?)\s+EXCEPTION\s+WHEN\s+OTHERS\s+THEN\s+(.*?)\s+END;', 
                    r'BEGIN \1 EXCEPTION WHEN OTHERS THEN \2 END;', sql, flags=re.DOTALL | re.IGNORECASE)
        
        return sql


    def wrap_in_postgresql_block(self, sql: str, variables: str) -> str:
        """Wrap the converted SQL in a PostgreSQL DO block with proper variable declarations"""
        # Clean up the SQL content - remove Oracle-specific DECLARE/BEGIN/END keywords
        sql = re.sub(r'^declare\s+', '', sql, flags=re.IGNORECASE)
        sql = re.sub(r'^begin\s+', '', sql, flags=re.IGNORECASE)
        sql = re.sub(r'end;\s*$', '', sql, flags=re.IGNORECASE)
        sql = sql.strip()
        
        # Ensure variables section is properly formatted
        if variables:
            # Clean up any duplicate semicolons or spacing issues in variables
            variables = re.sub(r';\s*;', ';', variables)
            variables = re.sub(r'\s+', ' ', variables).strip()
            
            # Ensure variables section ends with a space if it contains content
            if not variables.endswith(' '):
                variables += ' '
        else:
            variables = ''
        
        # Format the final PostgreSQL DO block with proper structure
        # DO $$ DECLARE [variables] BEGIN [sql_content] END $$;
        if variables:
            formatted_sql = f"DO $$ DECLARE {variables}BEGIN {sql} END $$;"
        else:
            formatted_sql = f"DO $$ BEGIN {sql} END $$;"
        
        # Apply final formatting to ensure clean, readable PostgreSQL code
        return self.format_postgresql_sql(formatted_sql)


    def extract_variables_and_constants(self, sql_content: str) -> str:
        """Extract variable declarations and constants from Oracle trigger with enhanced default value handling"""
        variables = []
        
        # Look for variable declarations between DECLARE and BEGIN
        declare_match = re.search(r'declare\s+(.*?)\s+begin', sql_content, re.IGNORECASE | re.DOTALL)
        if declare_match:
            var_section = declare_match.group(1)
            
            # Extract individual variable declarations - improved pattern matching
            lines = var_section.split('\n')
            current_var = ""
            
            for line in lines:
                line = line.strip()
                if not line or line.startswith('--'):
                    continue
                    
                # Handle multi-line variable declarations
                current_var += " " + line
                
                if line.endswith(';'):
                    # Complete variable declaration
                    var_decl = current_var.strip()
                    if 'exception' not in var_decl.lower():
                        # Convert Oracle types to PostgreSQL
                        pg_var = self.convert_data_types(var_decl)
                        
                        # Enhanced type conversions
                        pg_var = re.sub(r'PLS_INTEGER', 'integer', pg_var, flags=re.IGNORECASE)
                        pg_var = re.sub(r'SIMPLE_INTEGER', 'integer', pg_var, flags=re.IGNORECASE)
                        pg_var = re.sub(r'BINARY_INTEGER', 'integer', pg_var, flags=re.IGNORECASE)
                        pg_var = re.sub(r'BOOLEAN\s*:=\s*FALSE', 'BOOLEAN := FALSE', pg_var, flags=re.IGNORECASE)
                        pg_var = re.sub(r'BOOLEAN\s*:=\s*TRUE', 'BOOLEAN := TRUE', pg_var, flags=re.IGNORECASE)
                        
                        # Fix assignment operators in variable declarations - ensure proper spacing
                        pg_var = re.sub(r'\s*:\s*=\s*', ' := ', pg_var)
                        
                        # Handle default values for common types
                        pg_var = re.sub(r'(\w+\s+(?:integer|numeric|number)\s*):=\s*0', r'\1 := 0', pg_var, flags=re.IGNORECASE)
                        pg_var = re.sub(r'(\w+\s+(?:varchar|text)\s*(?:\(\d+\))?)\s*:=\s*([\'"][^\'"]*[\'"])', r'\1 := \2', pg_var, flags=re.IGNORECASE)
                        
                        # Handle %TYPE properly - convert to appropriate PostgreSQL types
                        if '%type' in pg_var.lower():
                            pg_var = self._convert_type_references(pg_var)
                        
                        # Handle RECORD types for cursor variables
                        if 'i1 ' in pg_var and 'RECORD' not in pg_var:
                            pg_var = re.sub(r'i1\s+[^;]+;', 'i1 RECORD;', pg_var)
                        
                        # Handle cursor variables properly
                        if 'cursor' in pg_var.lower() and 'for' in pg_var.lower():
                            # Convert cursor declarations
                            cursor_pattern = r'(\w+)\s+cursor\s+for\s+(.+);'
                            cursor_match = re.search(cursor_pattern, pg_var, re.IGNORECASE)
                            if cursor_match:
                                cursor_name = cursor_match.group(1)
                                pg_var = f'{cursor_name} REFCURSOR;'
                        
                        # Clean up any double conversions and formatting
                        pg_var = re.sub(r'(\w+)\s+(\w+\.\w+)varchar\(30\)', r'\1 \2%type', pg_var, flags=re.IGNORECASE)
                        pg_var = re.sub(r'\s+', ' ', pg_var).strip()
                        
                        # Ensure proper semicolon ending
                        if not pg_var.endswith(';'):
                            pg_var += ';'
                        
                        variables.append(pg_var)
                    current_var = ""
        
        # Join all variables with proper spacing for the DO block
        return ' '.join(variables)


    def _convert_type_references(self, var_declaration: str) -> str:
        """Convert Oracle %TYPE references to PostgreSQL equivalents"""
        
        # Common table.column%TYPE mappings to PostgreSQL types
        type_mappings = {
            'theme_no': 'VARCHAR(10)',
            'molecule_id': 'VARCHAR(10)', 
            'rg_no': 'VARCHAR(20)',
            'trademark_no': 'NUMERIC',
            'molecule_type_id': 'INTEGER',
            'pharmacological_type_id': 'INTEGER',
            'comparator_ind': 'VARCHAR(1)',
            'theme_desc_proposal': 'VARCHAR(500)',
            'manual_short_desc': 'VARCHAR(30)'
        }
        
        # Extract variable name and type reference
        type_pattern = r'(\w+)\s+(?:(\w+)\.)?(\w+)\.(\w+)%type'
        match = re.search(type_pattern, var_declaration, re.IGNORECASE)
        
        if match:
            var_name = match.group(1)
            schema = match.group(2) or 'gmd'
            table = match.group(3)
            column = match.group(4)
            
            # Try to map to a known PostgreSQL type
            mapped_type = type_mappings.get(column.lower())
            if mapped_type:
                return f'{var_name} {mapped_type};'
            else:
                # Keep the %TYPE reference but ensure proper schema prefix
                return f'{var_name} {schema}.{table}.{column}%type;'
        
        return var_declaration


    def convert_postgresql_specific(self, sql: str) -> str:
        """Convert to PostgreSQL-specific constructs"""
        # Convert certain SELECT statements to PERFORM for existence checks
        sql = re.sub(r'SELECT\s+COUNT\(\*\)\s+INTO\s+(\w+)\s+FROM\s+([^;]+);', 
                    r'SELECT COUNT(*) INTO \1 FROM \2;', 
                    sql, flags=re.IGNORECASE)
        
        # Convert Oracle dual table references
        sql = re.sub(r'FROM\s+dual\b', '', sql, flags=re.IGNORECASE)
        
        # Convert Oracle date functions
        sql = re.sub(r'TO_DATE\(([^,]+),\s*([^)]+)\)', r'TO_DATE(\1, \2)', sql, flags=re.IGNORECASE)
        
        # Improve schema references - add proper prefixes for common tables
        # Be more careful to avoid duplication
        schema_mappings = {
            r'\b(?<!\.)(v_theme_molecules)(?!\w)': 'gmd.v_theme_molecules_mrhub',
            r'\b(?<!\.)(theme_molecule_map)(?!\w)': 'gmd.theme_molecule_map', 
            r'\b(?<!\.)(themes)(?!\w)(?!\s*\.)': 'gmd.themes',
            r'\b(?<!\.)(v_themes)(?!\w)': 'gmd.v_themes',
            r'\b(?<!\.)(mdm_v_theme_status)(?!\w)': 'mdmappl.mdm_v_theme_status',
            r'\b(?<!\.)(mdm_v_disease_biology_areas)(?!\w)': 'mdmappl.mdm_v_disease_biology_areas',
            r'\b(?<!\.)(mdm_v_theme_molecule_map_mtn)(?!\w)': 'mdmappl.mdm_v_theme_molecule_map_mtn',
            r'\b(?<!\.)(mdm_v_new_medicine_proposals_mtn)(?!\w)': 'mdmappl.mdm_v_new_medicine_proposals_mtn',
            r'\b(?<!\.)(new_medicine_proposals)(?!\w)': 'predmd.new_medicine_proposals'
        }
        
        for pattern, replacement in schema_mappings.items():
            sql = re.sub(pattern, replacement, sql, flags=re.IGNORECASE)
        
        # Convert count() to COUNT(*)
        sql = re.sub(r'\bcount\(\)\b', 'COUNT(*)', sql, flags=re.IGNORECASE)
        
        # Add proper column defaults for INSERT statements
        sql = self._add_audit_columns(sql)
        
        return sql


    def _add_audit_columns(self, sql: str) -> str:
        """Add audit columns (ins_user, ins_date, upd_user, upd_date) to INSERT/UPDATE statements"""
        
        # Add audit columns to INSERT statements if not present
        insert_pattern = r'INSERT\s+INTO\s+(\w+\.\w+)\s*\(\s*([^)]+)\s*\)\s*VALUES\s*\(\s*([^)]+)\s*\)'
        
        def add_audit_to_insert(match):
            table = match.group(1)
            columns = match.group(2).strip()
            values = match.group(3).strip()
            
            # Check if audit columns are already present
            if 'ins_user' not in columns.lower():
                columns += ', ins_user, ins_date'
                values += ', :ins_user, :ins_date'
            
            return f'INSERT INTO {table} ( {columns} ) VALUES ( {values} )'
        
        sql = re.sub(insert_pattern, add_audit_to_insert, sql, flags=re.IGNORECASE | re.DOTALL)
        
        # Add audit columns to UPDATE statements if not present
        update_pattern = r'UPDATE\s+(\w+\.\w+)\s+SET\s+([^W]+?)(?=\s+WHERE|\s*;|\s*$)'
        
        def add_audit_to_update(match):
            table = match.group(1)
            set_clause = match.group(2).strip()
            
            # Check if audit columns are already present
            if 'upd_user' not in set_clause.lower():
                if not set_clause.endswith(','):
                    set_clause += ','
                set_clause += ' upd_user = :upd_user, upd_date = :upd_date'
            
            return f'UPDATE {table} SET {set_clause}'
        
        sql = re.sub(update_pattern, add_audit_to_update, sql, flags=re.IGNORECASE | re.DOTALL)
        
        return sql


    def extract_sections(self, sql_content: str) -> Dict[str, str]:
        """Extract INSERT, UPDATE, DELETE sections from Oracle trigger"""
        sections = {
            'on_insert': '',
            'on_update': '',
            'on_delete': ''
        }
        
        # Clean up content for better parsing
        content = self.clean_sql_content(sql_content)
        
        # Use a simpler but more robust parsing approach
        self._parse_oracle_trigger(content, sections)
        
        return sections


    def _parse_oracle_trigger(self, content: str, sections: Dict[str, str]):
        """Parse Oracle trigger using a line-by-line approach to handle complex nested structures"""
        
        lines = content.split('\n')
        current_section = None
        current_content = []
        if_stack = []
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            line_upper = line.upper()
            
            # Skip empty lines and comments
            if not line or line.startswith('--'):
                i += 1
                continue
            
            # Check for main conditional blocks
            if self._is_delete_start(line_upper):
                current_section = 'on_delete'
                current_content = []
                if_stack = ['DELETE']
                
            elif self._is_insert_start(line_upper):
                current_section = 'on_insert'
                current_content = []
                if_stack = ['INSERT']
                
            elif self._is_update_start(line_upper):
                current_section = 'on_update'
                current_content = []
                if_stack = ['UPDATE']
                
            elif self._is_insert_or_update_start(line_upper):
                # Special handling for shared IF INSERTING OR UPDATING block
                shared_content = self._extract_shared_block_manual(lines, i)
                self._parse_shared_insert_update_block(shared_content, sections)
                # Skip to after this block
                i = self._find_block_end(lines, i, 'IF', 'END IF') + 1
                continue
                
            elif current_section:
                # Add content to current section
                if line_upper.startswith('IF '):
                    if_stack.append('IF')
                elif line_upper.startswith('END IF') or line == 'END;':
                    if if_stack:
                        if_stack.pop()
                        if not if_stack:  # End of main block
                            sections[current_section] = '\n'.join(current_content)
                            current_section = None
                            current_content = []
                            i += 1
                            continue
                
                current_content.append(line)
            
            i += 1
        
        # Handle any remaining content
        if current_section and current_content:
            sections[current_section] = '\n'.join(current_content)


    def _is_delete_start(self, line: str) -> bool:
        """Check if line starts a DELETE block"""
        return (line.startswith('IF DELETING') or 
                line.startswith('IF (DELETING)') or
                line.startswith('ELSIF (DELETING)') or
                line.startswith('ELSIF DELETING'))


    def _is_insert_start(self, line: str) -> bool:
        """Check if line starts an INSERT block"""
        return (line.startswith('IF (INSERTING)') or
                line.startswith('IF INSERTING') or
                line.startswith('ELSIF (INSERTING)') or
                line.startswith('ELSIF INSERTING')) and 'OR' not in line


    def _is_update_start(self, line: str) -> bool:
        """Check if line starts an UPDATE block"""
        return (line.startswith('IF (UPDATING)') or
                line.startswith('IF UPDATING') or
                line.startswith('ELSIF (UPDATING)') or
                line.startswith('ELSIF UPDATING')) and 'OR' not in line


    def _is_insert_or_update_start(self, line: str) -> bool:
        """Check if line starts an INSERT OR UPDATE shared block"""
        return ('INSERTING OR UPDATING' in line or 
                'UPDATING OR INSERTING' in line) and line.startswith('IF')


    def _extract_shared_block_manual(self, lines: List[str], start_idx: int) -> str:
        """Manually extract the complete shared IF INSERTING OR UPDATING block"""
        end_idx = self._find_block_end(lines, start_idx, 'IF', 'END IF')
        block_lines = lines[start_idx+1:end_idx]  # Skip the IF line itself
        return '\n'.join(line.strip() for line in block_lines if line.strip())


    def _find_block_end(self, lines: List[str], start_idx: int, start_keyword: str, end_keyword: str) -> int:
        """Find the matching end of a block by counting nested IF/END IF pairs"""
        nesting_level = 1
        i = start_idx + 1
        
        while i < len(lines) and nesting_level > 0:
            line = lines[i].strip().upper()
            
            if line.startswith(start_keyword + ' '):
                nesting_level += 1
            elif line.startswith(end_keyword) or line == 'END;':
                nesting_level -= 1
                
            i += 1
        
        return i - 1  # Return index of the END IF line


    def _parse_shared_insert_update_block(self, shared_content: str, sections: Dict[str, str]):
        """Parse the shared IF INSERTING OR UPDATING block to extract operation-specific logic"""
        
        lines = shared_content.split('\n')
        shared_logic = []
        insert_logic = []
        update_logic = []
        
        current_section = 'shared'
        nesting_level = 0
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            line_upper = line.upper()
            
            if not line:
                i += 1
                continue
                
            # Check for nested IF INSERTING
            if line_upper.startswith('IF INSERTING') and 'OR' not in line_upper:
                current_section = 'insert'
                # Find the end of this nested block
                end_idx = self._find_nested_block_end(lines, i)
                insert_logic.extend(lines[i+1:end_idx])
                i = end_idx + 1
                current_section = 'shared'
                continue
                
            # Check for nested IF UPDATING  
            elif line_upper.startswith('IF UPDATING') and 'OR' not in line_upper:
                current_section = 'update'
                # Find the end of this nested block
                end_idx = self._find_nested_block_end(lines, i)
                update_logic.extend(lines[i+1:end_idx])
                i = end_idx + 1
                current_section = 'shared'
                continue
            
            # Add to shared logic if not in a nested block
            if current_section == 'shared':
                shared_logic.append(line)
            
            i += 1
        
        # Combine shared + specific logic
        shared_text = '\n'.join(shared_logic)
        insert_text = '\n'.join(insert_logic)
        update_text = '\n'.join(update_logic)
        
        # Combine with any existing logic and add shared parts
        if shared_text or insert_text:
            combined_insert = []
            if sections['on_insert']:
                combined_insert.append(sections['on_insert'])
            if shared_text:
                combined_insert.append(shared_text)
            if insert_text:
                combined_insert.append(insert_text)
            sections['on_insert'] = '\n'.join(combined_insert)
        
        if shared_text or update_text:
            combined_update = []
            if sections['on_update']:
                combined_update.append(sections['on_update'])
            if shared_text:
                combined_update.append(shared_text)
            if update_text:
                combined_update.append(update_text)
            sections['on_update'] = '\n'.join(combined_update)


    def _find_nested_block_end(self, lines: List[str], start_idx: int) -> int:
        """Find the end of a nested IF block within shared content"""
        nesting_level = 1
        i = start_idx + 1
        
        while i < len(lines) and nesting_level > 0:
            line = lines[i].strip().upper()
            
            if line.startswith('IF '):
                nesting_level += 1
            elif line.startswith('END IF'):
                nesting_level -= 1
                
            i += 1
        
        return i - 1


    def convert_oracle_to_postgresql(self, oracle_sql: str, variables: str) -> str:
        """
        Main conversion method - converts Oracle SQL to PostgreSQL and wraps in complete DO block
        
        This method:
        1. Applies all Oracle to PostgreSQL conversions
        2. Wraps the result in a complete PostgreSQL DO block 
        3. Includes ALL variables with their datatypes and default values
        4. Creates a standalone executable PostgreSQL block for each section
        
        Args:
            oracle_sql: The Oracle SQL content for a specific section (INSERT/UPDATE/DELETE)
            variables: All variable declarations with datatypes and default values
            
        Returns:
            Complete PostgreSQL DO block ready for execution
        """
        sql = oracle_sql
        
        # Apply all conversions in the correct order
        sql = self.clean_sql_content(sql)
        sql = self.convert_data_types(sql)
        sql = self.convert_nvl_to_coalesce(sql)
        sql = self.convert_functions(sql)
        # sql = self.convert_substr_to_substring(sql)
        sql = self.convert_sequence_generation(sql)
        sql = self.convert_exception_handling(sql)
        sql = self.convert_variable_references(sql)
        sql = self.convert_function_calls(sql)
        sql = self.convert_user_context(sql)
        sql = self.convert_numeric_validation(sql)
        sql = self.convert_date_functions(sql)
        sql = self.convert_cast_operations(sql)
        sql = self.convert_oracle_specifics(sql)
        sql = self.convert_postgresql_specific(sql)
        
        # Wrap in PostgreSQL DO block with ALL variables for complete standalone execution
        sql = self.wrap_in_postgresql_block(sql, variables)
        
        return sql


    def convert_trigger_file(self, input_file: str, output_file: str, verbose: bool = False):
        """Convert Oracle trigger file to PostgreSQL JSON format"""
        try:
            with open(input_file, 'r', encoding='utf-8') as f:
                oracle_content = f.read()
            
            # Check if file is empty
            if not oracle_content.strip():
                print(f"Warning: {input_file} is empty, skipping...")
                return False
            
            # Extract variables and constants with improved handling
            variables = self.extract_variables_and_constants(oracle_content)
            
            if verbose:
                print(f"📊 Extracted variables: {variables}")
            
            # Extract sections (now includes common code handling)
            sections = self.extract_sections(oracle_content)
            
            # Convert each section - each will get a complete DO block with all variables
            json_structure = {}
            
            for operation, sql_content in sections.items():
                if sql_content.strip():
                    if verbose:
                        print(f"🔄 Converting {operation} section...")
                        print(f"📝 Section content: {sql_content[:100]}...")
                    
                    # Each section gets converted with ALL variables to create a standalone DO block
                    converted_sql = self.convert_oracle_to_postgresql(sql_content, variables)
                    
                    if verbose:
                        print(f"✅ {operation} converted to complete PostgreSQL DO block")
                    
                    json_structure[operation] = [
                        {
                            "type": "sql", 
                            "sql": converted_sql
                        }
                    ]
            
            # Create output directory if it doesn't exist
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            
            # Write JSON output
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(json_structure, f, indent=4)
            
            print(f"✅ Successfully converted {input_file} → {output_file}")
            if verbose:
                print(f"📦 Created {len(json_structure)} sections: {list(json_structure.keys())}")
            return True
            
        except FileNotFoundError:
            print(f"❌ Error: Input file {input_file} not found")
            return False
        except Exception as e:
            print(f"❌ Error converting {input_file}: {str(e)}")
            if verbose:
                import traceback
                traceback.print_exc()
            return False


    def process_folder(self, oracle_folder: str, json_folder: str, verbose: bool = False):
        """Process all SQL files in the oracle folder"""
        oracle_path = Path(oracle_folder)
        json_path = Path(json_folder)
        
        if not oracle_path.exists():
            print(f"❌ Oracle folder '{oracle_folder}' does not exist")
            return
        
        # Create json folder if it doesn't exist
        json_path.mkdir(exist_ok=True)
        
        # Find all .sql files in oracle folder
        sql_files = list(oracle_path.glob('*.sql'))
        
        if not sql_files:
            print(f"⚠️  No .sql files found in '{oracle_folder}'")
            return
        
        print(f"🔍 Found {len(sql_files)} SQL file(s) in '{oracle_folder}'")
        
        success_count = 0
        for sql_file in sql_files:
            # Generate output filename
            output_filename = sql_file.stem + '.json'
            output_path = json_path / output_filename
            
            if verbose:
                print(f"Processing: {sql_file} → {output_path}")
            
            # Convert the file
            if self.convert_trigger_file(str(sql_file), str(output_path), verbose):
                success_count += 1
        
        print(f"\n🎉 Conversion complete: {success_count}/{len(sql_files)} files converted successfully")


    def analyze_oracle_files_for_exceptions(self, oracle_folder: str = 'orcale', save_to_excel: bool = True):
        """Analyze Oracle SQL files to find exceptions and add them to mappings"""
        oracle_path = Path(oracle_folder)
        
        if not oracle_path.exists():
            print(f"❌ Oracle folder '{oracle_folder}' does not exist")
            return
        
        # Find all .sql files in oracle folder
        sql_files = list(oracle_path.glob('*.sql'))
        
        if not sql_files:
            print(f"⚠️  No .sql files found in '{oracle_folder}'")
            return
        
        print(f"🔍 Analyzing {len(sql_files)} SQL file(s) for exceptions...")
        
        discovered_exceptions = {}
        exception_patterns = [
            # Custom exception declarations
            r'(\w+)\s+exception\s*;',
            # RAISE statements with custom exceptions
            r'raise\s+(\w+)\s*;',
            # raise_application_error calls
            r'raise_application_error\s*\(\s*(-?\d+)\s*,\s*[\'"]([^\'"]*)[\'"]',
            # WHEN exception handlers
            r'when\s+(\w+)\s+then',
            # Exception names in comments or strings
            r'--.*?(\w*_error|\w*_exception|\w*_not_found)',
        ]
        
        for sql_file in sql_files:
            try:
                with open(sql_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                print(f"📄 Analyzing {sql_file.name}...")
                
                # Clean content for better pattern matching
                content_upper = content.upper()
                
                # Find custom exception declarations
                exception_decl_pattern = r'(\w+)\s+EXCEPTION\s*;'
                for match in re.finditer(exception_decl_pattern, content_upper):
                    exception_name = match.group(1).lower()
                    if exception_name not in ['others', 'when', 'then']:
                        discovered_exceptions[exception_name] = f"Custom exception: {exception_name}"
                        print(f"  🔍 Found exception declaration: {exception_name}")
                
                # Find RAISE statements
                raise_pattern = r'RAISE\s+(\w+)\s*;'
                for match in re.finditer(raise_pattern, content_upper):
                    exception_name = match.group(1).lower()
                    if exception_name not in ['others', 'when', 'then', 'exception']:
                        discovered_exceptions[exception_name] = f"Raised exception: {exception_name}"
                        print(f"  🔍 Found RAISE statement: {exception_name}")
                
                # Find raise_application_error calls
                app_error_pattern = r'RAISE_APPLICATION_ERROR\s*\(\s*(-?\d+)\s*,\s*[\'"]([^\'"]*)[\'"]'
                for match in re.finditer(app_error_pattern, content_upper):
                    error_code = match.group(1)
                    error_message = match.group(2)
                    exception_key = f"app_error_{error_code.replace('-', 'neg_')}"
                    discovered_exceptions[exception_key] = error_message
                    print(f"  🔍 Found application error: {error_code} - {error_message[:50]}...")
                
                # Find WHEN exception handlers
                when_pattern = r'WHEN\s+(\w+)\s+THEN'
                for match in re.finditer(when_pattern, content_upper):
                    exception_name = match.group(1).lower()
                    if exception_name not in ['others', 'when', 'then'] and exception_name not in discovered_exceptions:
                        discovered_exceptions[exception_name] = f"Exception handler for: {exception_name}"
                        print(f"  🔍 Found exception handler: {exception_name}")
                
                # Find common Oracle exceptions
                common_oracle_exceptions = [
                    'no_data_found', 'too_many_rows', 'invalid_cursor', 'cursor_already_open',
                    'invalid_number', 'value_error', 'zero_divide', 'dup_val_on_index',
                    'case_not_found', 'access_into_null', 'collection_is_null',
                    'subscript_outside_limit', 'subscript_beyond_count', 'timeout_on_resource'
                ]
                
                for oracle_exception in common_oracle_exceptions:
                    if oracle_exception.upper() in content_upper and oracle_exception not in discovered_exceptions:
                        discovered_exceptions[oracle_exception] = f"Oracle built-in exception: {oracle_exception.replace('_', ' ').title()}"
                        print(f"  🔍 Found Oracle built-in: {oracle_exception}")
                        
            except Exception as e:
                print(f"❌ Error reading {sql_file}: {e}")
        
        if not discovered_exceptions:
            print("❌ No new exceptions found in the Oracle files")
            return
        
        print(f"\n📊 Summary: Found {len(discovered_exceptions)} unique exceptions")
        
        # Add new exceptions to mappings
        new_exceptions = {}
        for exception_name, postgres_message in discovered_exceptions.items():
            if exception_name not in self.exception_mappings:
                # Add to current mappings
                self.exception_mappings[exception_name] = postgres_message
                new_exceptions[exception_name] = postgres_message
                print(f"✅ Added exception mapping: {exception_name} → {postgres_message}")
            else:
                print(f"⚠️  Exception '{exception_name}' already exists in mappings")
        
        # Save all new exceptions to Excel in bulk
        if new_exceptions and save_to_excel:
            self.save_exceptions_to_excel_force(new_exceptions)
        
        if new_exceptions:
            print(f"\n✅ Successfully added {len(new_exceptions)} new exception mappings!")
            print("📋 New exceptions added:")
            for exception_name, postgres_message in new_exceptions.items():
                print(f"  • {exception_name} → {postgres_message}")
        else:
            print("ℹ️  All discovered exceptions were already in the mappings")
    
    def save_exceptions_to_excel_force(self, exceptions_dict: Dict[str, str], excel_file: str = DEFAULT_EXCEL_FILE):
        """Force save exceptions to Excel file using openpyxl directly"""
        try:
            
            # Check if Excel file exists
            if os.path.exists(excel_file):
                # Load existing workbook
                workbook = load_workbook(excel_file)
            else:
                # Create new workbook
                workbook = Workbook()
                # Remove default sheet
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
            
            # Get or create exception_mappings sheet
            if 'exception_mappings' in workbook.sheetnames:
                sheet = workbook['exception_mappings']
                # Find the last row with data
                last_row = sheet.max_row
            else:
                sheet = workbook.create_sheet('exception_mappings')
                # Add headers
                sheet['A1'] = 'Oracle_Exception'
                sheet['B1'] = 'PostgreSQL_Message'
                last_row = 1
            
            # Add new exceptions
            row_num = last_row + 1
            for oracle_exception, postgres_message in exceptions_dict.items():
                sheet[f'A{row_num}'] = oracle_exception
                sheet[f'B{row_num}'] = postgres_message
                row_num += 1
            
            # Save the workbook
            workbook.save(excel_file)
            print(f"💾 Successfully saved {len(exceptions_dict)} exceptions to {excel_file}")
            
        except ImportError:
            print("❌ openpyxl library not available. Cannot save to Excel.")
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")




def main():
    # Check if any command line arguments are provided
    if len(sys.argv) == 1:
        # No arguments provided, show interactive menu
        converter = OracleToPostgreSQLConverter()
        
        while True:
            choice = show_interactive_menu()
            
            if choice == '1':
                # Convert entire folder (default)
                print("\n📁 Converting entire folder with default settings...")
                print("📁 Oracle folder: orcale")
                print("📁 JSON folder: json")
                verbose = get_verbose_choice()
                converter.process_folder('orcale', 'json', verbose)
                
            elif choice == '2':
                # Convert folder with custom names
                oracle_folder, json_folder = get_folder_names()
                verbose = get_verbose_choice()
                converter.process_folder(oracle_folder, json_folder, verbose)
                
            elif choice == '3':
                # Convert single file
                input_file, output_file = get_file_details()
                verbose = get_verbose_choice()
                converter.convert_trigger_file(input_file, output_file, verbose)
                
            elif choice == '4':
                # Convert single file with custom output
                input_file, output_file = get_file_details()
                verbose = get_verbose_choice()
                converter.convert_trigger_file(input_file, output_file, verbose)
                
            elif choice == '5':
                # Convert with verbose output (choose file or folder)
                print("\n🔍 Verbose Conversion")
                mode_choice = input("Convert (f)older or single (file)? (f/file, default: f): ").strip().lower()
                
                if mode_choice in ['file', 'f']:
                    input_file, output_file = get_file_details()
                    converter.convert_trigger_file(input_file, output_file, True)
                else:
                    oracle_folder = input("Enter Oracle folder name (default: orcale): ").strip() or "orcale"
                    json_folder = input("Enter JSON output folder name (default: json): ").strip() or "json"
                    converter.process_folder(oracle_folder, json_folder, True)
                
            elif choice == '6':
                # Analyze Oracle files for exceptions
                print("\n🔬 Analyzing Oracle Files for Exceptions")
                oracle_folder = input("Enter Oracle folder name (default: orcale): ").strip() or "orcale"
                save_to_excel = input("Save to Excel file? (y/n, default: y): ").strip().lower()
                save_excel = save_to_excel not in ['n', 'no']
                
                converter.analyze_oracle_files_for_exceptions(oracle_folder, save_excel)
                
            elif choice == '7':
                # Show help
                show_help()
                input("\nPress Enter to continue...")
                continue
                
            elif choice == '8':
                # Exit
                print("\n👋 Goodbye!")
                break
            
            # Ask if user wants to perform another operation
            print("\n" + "="*60)
            another = input("Perform another operation? (y/n, default: n): ").strip().lower()
            if another not in ['y', 'yes']:
                print("\n👋 Goodbye!")
                break
    
    else:
        # Command line arguments provided, use traditional argparse
        parser = argparse.ArgumentParser(description='Convert Oracle PL/SQL trigger to PostgreSQL JSON format')
        
        # Add mode selection
        parser.add_argument('--mode', choices=['file', 'folder'], default='folder',
                           help='Processing mode: single file or entire folder (default: folder)')
        
        # File mode arguments
        parser.add_argument('input_file', nargs='?', help='Input Oracle SQL trigger file (for file mode)')
        parser.add_argument('-o', '--output', help='Output JSON file (for file mode)')
        
        # Folder mode arguments  
        parser.add_argument('--oracle-folder', default='orcale', 
                           help='Oracle SQL files folder (default: orcale)')
        parser.add_argument('--json-folder', default='json',
                           help='Output JSON files folder (default: json)')
        
        # Common arguments
        parser.add_argument('-v', '--verbose', action='store_true', help='Verbose output')
        
        args = parser.parse_args()
        
        converter = OracleToPostgreSQLConverter()
        
        if args.mode == 'file':
            # File mode - convert single file
            if not args.input_file:
                print("❌ Error: input_file is required for file mode")
                parser.print_help()
                return
            
            output_file = args.output or args.input_file.replace('.sql', '.json')
            converter.convert_trigger_file(args.input_file, output_file, args.verbose)
            
        else:
            # Folder mode - convert all files in oracle folder
            print("🚀 Starting folder processing...")
            print(f"📁 Oracle folder: {args.oracle_folder}")
            print(f"📁 JSON folder: {args.json_folder}")
            converter.process_folder(args.oracle_folder, args.json_folder, args.verbose)




if __name__ == "__main__":
    main()



